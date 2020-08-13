from bs4 import BeautifulSoup
import requests, re, unidecode, math, time

###################### Setup #########################

# Team Codes
epl_team_codes = {'Liverpool': "822bd0ba", 'Leicester City': "a2d435b3", 'Manchester City': "b8fd03ef", 'Chelsea': "cff3d9bb", 'Wolves': "8cec06e1", 'Sheffield Utd': "1df6b87e", 'Burnley': "943e8050", 'Arsenal': "18bb7c10", 'Manchester Utd': "19538871", 'Tottenham': "361ca564", 'Bournemouth': "4ba7cbea", 'Brighton': "d07537b9", 'Crystal Palace': "47c64c55", 'Newcastle Utd': "b2b47a98", 'Aston Villa': "8602292d", 'Everton': "d3fd31cc", 'West Ham': "7c21e445", 'Norwich City': "1c781004", 'Southampton': "33c895d4", 'Watford': "2abfe087"}
la_liga_team_codes = {'Barcelona': "206d90db", 'Real Madrid': "53a2f082", 'Sevilla': "ad2be733", 'Atletico Madrid': "db3b9613", 'Real Sociedad': "e31d1cd9", 'Getafe': "7848bd64", 'Athletic Bilbao': "2b390eca", 'Valencia': "dcc91a7b", 'Levante': "9800b6a1", 'Villareal': "2a8183b3", 'Granada': "a0435291", 'Osasuna': "03c57e2b", 'Betis': "fc536746", 'Valladolid': "17859612", 'Alaves': "8d6fd021", 'Eibar': "bea5c710", 'Mallorca': "2aa12281", 'Celta Vigo': "f25da7fb", 'Leganes': "7c6f2c78", 'Espanyol': "a8661628"}
serie_a_team_codes = {'Juventus': "e0652b02", 'Inter': "d609edc0", 'Lazio': "7213da33", 'Roma': "cf74a709", 'Atalanta': "922493f3", 'Cagliari': "c4260e09", 'Parma': "eab4234c", 'Napoli': "d48ad4ff", 'Torino': "105360fe", 'Hellas Verona': "0e72edf2", 'Bologna': "1d8099f8", 'Milan': "dc56fe14", 'Sassuolo': "e2befd26", 'Udinese': "04eea015", 'Fiorentina': "421387cf", 'Lecce': "ffcbe334", 'Sampdoria': "8ff9e3b3", 'Genoa': "658bf2de", 'Brescia': "4ef57aeb", 'SPAL': "1d2fe027"}
championship_team_codes = {'Leeds United': "5bfb9659", 'West Brom': "60c6b05f", 'Brentford': "cd051869", 'Nottingham Forest': "e4a775cb", 'Fulham': "fd962109", 'Swansea City': "fb10988f", 'Millwall': "e3c537a1", 'Sheffield Weds': "bba7d733", 'Hull City': "bd8769d1", 'Preston': "22df8478", 'Bristol City': "93493607", 'Cardiff City': "75fae011", 'Blackburn': "e090f40b", 'Reading': "b0ac61ff", 'QPR': "a757999c", 'Middlesbrough': "7f59c601", 'Derby County': "26ab47ee", 'Birmingham City': "ec79b7c2", 'Charlton Athletic': "7a8db6d4", 'Huddersfield': "f5922ca5", 'Stoke City': "17892952", 'Wigan Athletic': "e59ddc76", 'Barnsley': "293cb36b", 'Luton Town': "e297cd13"}

league_codes = [epl_team_codes,la_liga_team_codes,serie_a_team_codes,championship_team_codes]

team_template = "https://fbref.com/en/squads/"

# Stat arrays
player = ""
position = ""
height = ""
weight = ""
dob = ""
footed = ""
dates = []
dows = []
comps = []
rounds =[]
venues = []
results = []
squads = []
opponents = []
games_started = []
minutes = []
goals = []
assists = []
pens = []
pen_attempts = []
total_shots = []
sots = []
yellows = []
reds = []
#removed crosses[]
touches = [] #NEW
pressures = [] #NEW
tackles = []
interceptions = []
blocks = [] #NEW
passes_completed = [] #NEW
passes_attempted = [] #NEW
passes_fwd_dist = [] #NEW
carries = [] #NEW
carries_fwd_dist = [] #NEW
dribbles_attempted = [] #NEW
dribbles_completed = [] #NEW
#removed fouls_committed = []
#removed fouls_suffered = []
xgs = []
npxgs = []
xas = []
sca = []
gca = []

######################################################

##################### Functions ######################

# Returns fantasy score given a player's stats
def get_fantasy_score(stats):
    scoring = [10, 6, 1, 1, .7, 1, 0, 0, 0, 0, 0, 1, .5, -.5, -1.5, -3]
    score = 0.00
    for index in range(len(stats)):
        try:
            score += round(float(stats[index]*scoring[index]),1)
        except:
            continue
    return score

# Returns standard deviation of a player's fantasy points
def get_std(fps):

    # Get average
    sum = 0
    for fp in fps:
        sum += fp
    average = round(float(sum/len(fps)),2)
    squared = []

    # Get squared average
    for fp in fps:
        squared.append(round(float((fp-average)**2),2))
    sum_squared = 0
    for square in squared:
        sum_squared += square
    squared_average = round(float(sum_squared/len(squared)),2)

    # Return square root of squared average
    return math.sqrt(squared_average)

# Returns stat types for a player
def get_stats(soup):
    stat_types = []
    data = []
    try:
        data = soup.findAll('td')
    except AttributeError:
        print("Cannot find player.")
    for element in data:
        stat = element.get('data-stat')
        if stat not in stat_types:
            stat_types.append(stat)
    return stat_types

# Returns formatted array of all stats for a player
def get_player_stats(pid):

    # Gets player's page
    player_page = "https://fbref.com/en/players/"+pid+"/matchlogs/2019-2020/-Match-Logs"
    html_player = requests.get(player_page).text

    # Soups player's webpage
    player_soup = BeautifulSoup(html_player, 'html.parser').find('div', attrs={'id': "all_ks_matchlogs_all"})
    if player_soup == None:
        player_soup = BeautifulSoup(html_player, 'html.parser')

    # Gets stat types and removes unnecessary ones
    stat_types = get_stats(player_soup)
    try:
        stat_types.remove("match_report")
    except ValueError:
        pass

    # Adds player name (using PID) for each stat
    player = "Matches"
    while player == "Matches":
        player = soup.find('a', attrs={'href': re.compile(pid)}).text
    player = unidecode.unidecode(player)
    player_message = "Gathering stats for %s..." % player
    print(player_message)

    #Gets player's personal metrics
    metrics_soup = BeautifulSoup(html_player, 'html.parser')
    #print(metrics_soup)
    try:
        position = metrics_soup.find(text = re.compile('Position:')).parent.next_sibling.strip()
        weird_things = [' ','\n','\t','▪']
        for thing in weird_things:
            position = position.replace(thing,'')
        position = position[:-1]
    except:
        position = ''
    try:
        height = metrics_soup.find('span', attrs={'itemprop': "height"}).text
        height = round(int(height[:-2].replace(' ',''))*.393701,1) # Height in inches
    except:
        height = ''
    try:
        weight = metrics_soup.find('span', attrs={'itemprop': "weight"}).text
        weight = round(int(weight[:-2].replace(' ',''))*2.20462,1) # Weight in pounds
    except:
        weight = ''
    try:
        dob = metrics_soup.find('span', attrs={'data-birth': re.compile("\d\d\d\d-\d+\d+")}).attrs['data-birth'].replace(' ','').replace('\t','')
    except:
        dob = ''
    try:
        footed = metrics_soup.find(text = re.compile('Footed:')).parent.next_sibling.strip()[3:]
    except:
        footed = ''
    #print(position,height,weight,dob,footed)

    # Adds dates (using Regex to find them)
    for element in player_soup.findAll('a', attrs={'href': re.compile("-\w+-\w+-\d+-\d\d\d\d-")}):
        if element.text != "Match Report":
            dates.append(element.text)

    # Assigns stats to correct array
    bench_dates = []
    for stat_type in stat_types:
        for stat in player_soup.findAll('td', attrs={'data-stat': stat_type}):
            if stat_type == 'dayofweek':
                dows.append(stat.text)
            elif stat_type == 'comp':
                comps.append(stat.text)
            elif stat_type == 'round':
                rounds.append(stat.text)
            elif stat_type == 'venue':
                venues.append(stat.text)
            elif stat_type == 'result':
                results.append(stat.text)
            elif stat_type == 'squad':
                text = stat.text
                first = text[:1]
                if first.islower():
                    spot = 1
                    for char in text:
                        if char.isspace():
                            text = text[spot:]
                            break
                        spot += 1
                squads.append(text)
            elif stat_type == 'opponent':
                text = stat.text
                first = text[:1]
                if first.islower():
                    spot = 1
                    for char in text:
                        if char.isspace():
                            text = text[spot:]
                            break
                        spot += 1
                opponents.append(text)
            elif stat_type == 'game_started':
                games_started.append(stat.text)
            elif stat_type == 'minutes':
                minutes.append(stat.text)
            elif stat_type == 'goals':
                goals.append(stat.text)
            elif stat_type == 'assists':
                assists.append(stat.text)
            elif stat_type == 'pens_made':
                pens.append(stat.text)
            elif stat_type == 'pens_att':
                pen_attempts.append(stat.text)
            elif stat_type == 'shots_total':
                total_shots.append(stat.text)
            elif stat_type == 'shots_on_target':
                sots.append(stat.text)
            elif stat_type == 'cards_yellow':
                yellows.append(stat.text)
            elif stat_type == 'cards_red':
                reds.append(stat.text)
            elif stat_type == 'touches':
                touches.append(stat.text)
            elif stat_type == 'pressures':
                pressures.append(stat.text)
            elif stat_type == 'tackles':
                tackles.append(stat.text)
            elif stat_type == 'interceptions':
                interceptions.append(stat.text)
            elif stat_type == 'blocks':
                blocks.append(stat.text)
            elif stat_type == 'passes_completed':
                passes_completed.append(stat.text)
            elif stat_type == 'passes':
                passes_attempted.append(stat.text)
            elif stat_type == 'passes_progressive_distance':
                passes_fwd_dist.append(stat.text)
            elif stat_type == 'carries':
                carries.append(stat.text)
            elif stat_type == 'carry_progressive_distance':
                carries_fwd_dist.append(stat.text)
            elif stat_type == 'dribbles_completed':
                dribbles_completed.append(stat.text)
            elif stat_type == 'dribbles':
                dribbles_attempted.append(stat.text)
            elif stat_type == 'xg':
                xgs.append(stat.text)
            elif stat_type == 'npxg':
                npxgs.append(stat.text)
            elif stat_type == 'xa':
                xas.append(stat.text)
            elif stat_type == 'sca':
                sca.append(stat.text)
            elif stat_type == 'gca':
                gca.append(stat.text)
            elif stat_type == 'bench_explain':
                for element in player_soup.findAll('td', {'data-stat':"bench_explain"}):
                    if element.text == "On matchday squad, but did not play":
                        parent_soup = element.parent
                        date_element = parent_soup.find('a', attrs={'href': re.compile("-\w+-\w+-\d+-\d\d\d\d-")})
                        bench_dates.append(date_element.text)

    # Gets unique dates
    bench_dates.sort()
    temp = []
    for date in bench_dates:
        if date not in temp:
            temp.append(date)
    bench_dates = temp

    # Removes extra lines for summary stats
    all_stats = [player, dates, dows, comps, rounds, venues, results, squads, opponents, games_started, minutes, goals, assists, pens, pen_attempts, total_shots, sots, yellows, reds, touches, pressures, tackles, interceptions, blocks, xgs, npxgs, xas, sca, gca, passes_completed, passes_attempted, passes_fwd_dist, carries, carries_fwd_dist, dribbles_completed, dribbles_attempted]
    for stat in all_stats:
        if stat != dates and stat != player and len(stat) > 0:
            del stat[len(stat)-1]

    # Space addition prep
    play_stats = [minutes, goals, assists, pens, pen_attempts, total_shots, sots, yellows, reds, touches, pressures, tackles, interceptions, blocks, xgs, npxgs, xas, sca, gca, passes_completed, passes_attempted, passes_fwd_dist, carries, carries_fwd_dist, dribbles_completed, dribbles_attempted]
    line_spaces = []
    didnt_play_spaces = []

    # Adds 0s to cup match xas, xgs, etc.
    for stat in play_stats:
        for datapoint in stat:
            if datapoint == '':
                stat[stat.index(datapoint)] = 0.0
            try:
                stat[stat.index(datapoint)] = float(datapoint) # Turns strings to floats
            except ValueError:
                try:
                    stat[stat.index(datapoint)] = int(datapoint) # Or turns string to ints
                except ValueError:
                    pass

    # Gets long line spaces
    for entry in range(len(rounds)):
        if rounds[entry] == '':
            line_spaces.append(entry)

    # Gets didn't play spaces
    for entry in bench_dates:
        didnt_play_spaces.append(dates.index(entry))

    # Sorts lists by order of occurrence
    line_spaces.sort()
    didnt_play_spaces.sort()

    # Adds appropriate spaces
    for stat in play_stats:
        if len(stat) > 0 and len(didnt_play_spaces) > 0:
            for space in didnt_play_spaces:
                try:
                    stat.insert(space, '')
                except IndexError:
                    stat.append('')
        if len(stat) > 0 and len(line_spaces) > 0:
            for space in line_spaces:
                try:
                    stat.insert(space, '')
                except IndexError:
                    stat.append('')

    # Adds line spaces to dates, which are weird
    for space in line_spaces:
        try:
            dates.insert(space, '')
        except IndexError:
            dates.append('')

    # Add metrics into appropriate spots
    metrics = [position, height, weight, dob, footed]
    spot = 1
    for metric in metrics:
        all_stats.insert(spot,metric)
        spot += 1

    # Returns stats
    return all_stats

######################################################

# Will hold teams/leagues to scrape
teams = []
leagues = []

# User input
print()
selection = input("Choose [TEAM], [LEAGUE], or [ALL]: ")
print()

# Validity to check if response was valid
valid = False
while not valid:
    if selection == "TEAM":
        valid = True
        teams.clear()
        team_found = False
        team_selection = input("Choose your team: ")
        while not team_found:
            for league in league_codes: # Checks to see if team is valid
                if team_selection in league.keys():
                        teams.append(team_selection)
                        leagues.append(league)
                        team_found = True
            if not team_found:
                print()
                team_selection = input("Team not found. Please try again: ")
    elif selection == "LEAGUE":
        valid = True
        leagues.clear()
        league_found = False
        which_league = input("Choose [EPL], [EPL2], [LIGA], or [SERIEA]: ")
        while not league_found:
            if which_league == "EPL":
                leagues.append(epl_team_codes)
                league_found = True
            elif which_league == "EPL2":
                leagues.append(championship_team_codes)
                league_found = True
            elif which_league == "LIGA":
                leagues.append(la_liga_team_codes)
                league_found = True
            elif which_league == "SERIEA":
                leagues.append(serie_a_team_codes)
                league_found = True
            else:
                print()
                which_league = input("League not found. Please try again: ")
    elif selection == "ALL":
        valid = True
        leagues.clear()
        for league in league_codes:
            leagues.append(league)
    else:
        selection = input("Invalid input. Please choose [TEAM], [LEAGUE], or [ALL]: ")
        print()

# To run one team vs. full league
if selection != 'TEAM':
    for league in leagues:
        for team in league.keys():
            teams.append(team)

# Excel file setup
try:
	from openpyxl import Workbook
	from openpyxl import load_workbook
	xlInstalled = True
except ImportError:
	print("Please install openpyxl if you want Excel output!")
	xlInstalled = False
if xlInstalled:
	excelFile = input("Excel file for output: ")
	if not excelFile.endswith(".xlsx") or not excelFile.endswith(".xls"):
		excelFile += ".xlsx"
	wb = Workbook()
	ws = wb.active
	ws.title = "%s Stats" % selection
	ws.append(["Player","Position","Height","Weight","DoB","Foot","Date","Day of Week","Round","Venue","Result","Squad","Opponent","Started","Minutes","Goals","Assists","PKs Scored","PKs Attempted","Shots","Shots on Target","Yellows","Reds","Touches","Pressures","Tackles","Interceptions","Blocks","xG","npxG","xA","Sca","Gca","Passes Completed","Passes Attempted","Fwd Pass Dist","Carries","Fwd Carry Distance","Dribbles Completed","Dribbles Attempted","Fantasy Points"])
print()

# Create new averages sheet
ws2 = wb.create_sheet("Averages",1)
ws2.title = "%s Averages" % selection
ws2.append(["Player","Squad","Position","Height","Weight","DoB","Foot","Total Matches","Minutes","Goals","Assists","Shots","Shots on Target","Crosses","Fouled","PKs Scored","PKs Attempted","xG","npxG","xA","Tackles","Interceptions","Fouls Committed","Yellows","Reds","Fantasy Points","Standard Deviation"])
num_team = 1
total_time = 0
for team in teams:
    start_time = time.time()
    if num_team == 1:
        average_time = 0
        time_string = "N/A"
    else:
        average_time = total_time/(num_team-1)
        estimated_time = int(round(average_time*(len(teams)-(num_team-1))))
        ss = estimated_time%60
        ms = int(round((estimated_time-ss)/60))
        time_string = str(ms)+"m "+str(ss)+"s"
    print("Welcome to "+team+"! ("+str(num_team)+"/"+str(len(teams))+") [Estimated time remaining: "+time_string+"]")
    print()

    # Requests team webpage based on input and converts to soup
    for league in leagues:
        if team in league.keys():
            r = requests.get(team_template+league[team])
            break
    html_text = r.text
    soup = BeautifulSoup(html_text, 'html.parser')

    # Gets all player ids
    pids_all = []
    for link in soup.findAll('a', attrs={'href': re.compile("/en/players/\w+/matchlogs/\d\d\d\d-\d\d\d\d/.*-Match-Logs")}):
        pids_all.append(link.get('href')[12:20])
    pids = []
    for id in pids_all:
        if id not in pids:
            pids.append(id)

    # For all players...
    players_completed = []
    for id in pids:

        # Fantasy point prep
        total_fps = 0
        fps_to_std = []

        # Empty player stat lists
        complete_stats = [player, dates, dows, comps, rounds, venues, results, squads, opponents, games_started, minutes, goals, assists, pens, pen_attempts, total_shots, sots, yellows, reds, touches, pressures, tackles, interceptions, blocks, xgs, npxgs, xas, sca, gca, passes_completed, passes_attempted, passes_fwd_dist, carries, carries_fwd_dist, dribbles_completed, dribbles_attempted]
        for array in complete_stats:
            try:
                del array[:]
            except:
                array = ""

        # Add new player stats to Excel sheet
        player_stats = get_player_stats(id)

        # Delete wrong competitions
        to_delete = []
        spot = 0
        for comp in player_stats[8]:
            if comp != "Premier League" and comp != "La Liga" and comp != "Serie A" and comp != "Champions Lg":
                if player_stats[9][spot][:9] != "Matchweek":
                    to_delete.append(spot)
            spot += 1
        del player_stats[8]
        for stat in player_stats[6:]: # All non-metrics
            new_stat = []
            for datapoint in range(len(stat)):
                if datapoint not in to_delete:
                    new_stat.append(stat[datapoint])
            player_stats[player_stats.index(stat)] = new_stat

        num_matches = len(player_stats[8]) # Rounds (but shouldn't matter too much)
        for entry in range(num_matches):
            match = []
            num_metrics = 6
            for spot in range(num_metrics):
                match.append(player_stats[spot])
            for stat in player_stats:
                if player_stats.index(stat) > 5: # If not the name...
                    try:
                        match.append(stat[entry])
                    except:
                        continue
            fp = get_fantasy_score(match[24:]) # Goals to red cards
            total_fps += fp
            fps_to_std.append(fp)
            match.append(fp)
            if team in championship_team_codes.keys():  # Championship sides don't have xg, npxg, or xa
                for time in range(3):
                    match.insert(23,'')
            try:
                if match[14] != '' and match[0] not in players_completed: # Don't append if stats are empty (i.e. didn't play or line) or player already done
                    ws.append(match)
            except:
                pass

        # Add average stats to average sheet
        averages = player_stats[:6] # Final array to add
        to_average = player_stats[24:] # Stats that need to be averaged
        averaged = []

        # Adding up all of a player's single stat and dividing by number of data points
        for group in to_average:
            total = 0
            for stat in group:
                try:
                    total += stat
                except:
                    continue
            try:
                averaged.append(round(float(total/len(group)),2)) # Adding averaged stat to final average array
            except:
                averaged.append('')

        # Making stats per 90
        try:
            averaged.append(round(float(total_fps/num_matches),2)) # Adds average fantasy points to end of average array
        except:
            pass
        min = averaged[0]
        try:
            coefficient = round(float(90/min),5) # Gets coefficient to multiply all stats by, which is just what it takes to get average minutes to 90
        except:
            coefficient = 0
        for stat in averaged:
            try:
                averaged[averaged.index(stat)] = round(stat*coefficient,2) # Multiplies each stat by "per 90" coefficient
            except:
                continue

        # Replace minutes with total min
        total_min = 0
        for min in player_stats[14]:
            try:
                total_min += min
            except:
                continue

        # Add standard deviation to end of average array
        try:
            std = get_std(fps_to_std)
            averaged.append(std)
        except:
            continue

        # Puts two arrays together and adds a couple other data points
        if num_matches > 0:
            averages.extend(averaged)
            averages.insert(1,player_stats[11][0]) # Squad
            averages.insert(7,num_matches)
            averages[8] = total_min
            if averages[8] != 0 and averages[0] not in players_completed:
                ws2.append(averages)

        players_completed.append(player_stats[0]) # So we don't have repeat goalkeepers in Excel file

    num_team+=1
    total_time += (time.time()-start_time)
    print()

# Save Excel file
if xlInstalled:
	wb.save(excelFile)

print()
print("Done!")
print()
